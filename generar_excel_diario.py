"""
Genera una copia actualizada de Estado_conexiones.xlsx con los valores más
recientes de Fecha / Hora / Status de TODOS los dispositivos que existen hoy
en ChirpStack.

A diferencia de la versión anterior, las filas ya no salen de un diccionario
fijo (APLICACIONES / ENTIDAD_A_DEVEUI): se descubren automáticamente todas
las Ubicaciones (aplicaciones) y Entidades (dispositivos) del tenant vía
chirpstack_client.descubrir_todo(). Si aparece una ubicación o un equipo
nuevo en ChirpStack, se agrega solo como fila nueva la próxima vez que corre
este script -- no hace falta editar código.

Los únicos ajustes manuales que se mantienen (equipos que comparten un mismo
DevEUI, o filas para equipos aún no dados de alta en ChirpStack) viven en
entidades_manuales.py.

No modifica ningún archivo en OneDrive/SharePoint: crea un archivo nuevo
(con la fecha del día en el nombre) que luego el workflow de GitHub Actions
adjunta a un correo.

Variables de entorno esperadas:
  CHIRPSTACK_API_TOKEN   -> token generado en ChirpStack (Tenant > API Keys)
"""

import os
from copy import copy
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from chirpstack_client import descubrir_todo, calcular_estado, TZ_CHILE
from entidades_manuales import DUPLICAR_ENTIDAD, FILAS_MANUALES, UBICACIONES_EXCLUIDAS
from historial import actualizar_historial, dias_ordenados

API_TOKEN = os.environ["CHIRPSTACK_API_TOKEN"]
PLANTILLA = "Estado_conexiones_template.xlsx"
HOJA = "Hoja1"
PRIMERA_FILA_DATOS = 3
COLUMNAS = 7  # A..G: Ubicación, Entidad, Fecha, Hora, Status, DevEUI, ID


def construir_filas(ahora=None):
    """Arma la lista completa de filas (Ubicación, Entidad, Fecha, Hora,
    Status, DevEUI) combinando lo descubierto en ChirpStack con los ajustes
    manuales de entidades_manuales.py."""
    filas = []
    for d in descubrir_todo(API_TOKEN):
        if d["ubicacion"] in UBICACIONES_EXCLUIDAS:
            continue
        fecha, hora, status = calcular_estado(d["last_seen"], ahora=ahora)
        nombres = DUPLICAR_ENTIDAD.get(d["dev_eui"]) or [d["entidad"]]
        for nombre in nombres:
            filas.append({
                "ubicacion": d["ubicacion"],
                "entidad": nombre,
                "fecha": fecha,
                "hora": hora,
                "status": status,
                "dev_eui": d["dev_eui"],
            })

    for extra in FILAS_MANUALES:
        if extra["ubicacion"] in UBICACIONES_EXCLUIDAS:
            continue
        filas.append({
            "ubicacion": extra["ubicacion"],
            "entidad": extra["entidad"],
            "fecha": None,
            "hora": None,
            "status": extra.get("status", "Pendiente"),
            "dev_eui": extra.get("dev_eui", ""),
        })

    filas.sort(key=lambda f: (f["ubicacion"], f["entidad"]))
    return filas


def escribir_excel(filas, plantilla=PLANTILLA):
    wb = openpyxl.load_workbook(plantilla)
    ws = wb[HOJA]

    # Estilo de referencia: en esta plantilla todas las filas de datos
    # comparten el mismo estilo, así que basta copiar el de la primera fila.
    estilo_ref = {
        col: copy(ws.cell(row=PRIMERA_FILA_DATOS, column=col)._style)
        for col in range(1, COLUMNAS + 1)
    }
    numfmt_fecha = ws.cell(row=PRIMERA_FILA_DATOS, column=3).number_format
    numfmt_hora = ws.cell(row=PRIMERA_FILA_DATOS, column=4).number_format

    # Limpia cualquier fila de datos vieja (por si la corrida anterior tenía
    # más o menos filas que esta -- ej. se dieron de baja equipos).
    ultima_fila = max(ws.max_row, PRIMERA_FILA_DATOS + len(filas)) + 5
    for row in range(PRIMERA_FILA_DATOS, ultima_fila + 1):
        for col in range(1, COLUMNAS + 1):
            ws.cell(row=row, column=col).value = None

    for i, f in enumerate(filas):
        row = PRIMERA_FILA_DATOS + i
        valores = [
            f["ubicacion"], f["entidad"], f["fecha"], f["hora"], f["status"],
            f["dev_eui"], f"{f['ubicacion']}|{f['entidad']}",
        ]
        for col, valor in enumerate(valores, start=1):
            celda = ws.cell(row=row, column=col)
            celda.value = valor
            celda._style = copy(estilo_ref[col])
        ws.cell(row=row, column=3).number_format = numfmt_fecha
        ws.cell(row=row, column=4).number_format = numfmt_hora

    hoy = datetime.now(TZ_CHILE).strftime("%Y-%m-%d")
    nombre_salida = f"Estado_conexiones_{hoy}.xlsx"
    wb.save(nombre_salida)
    return nombre_salida


# Colores de la hoja "Resumen": verde = Conectado, rojo = Desconectado,
# gris = sin dato para ese día (equipo pendiente, recién agregado, etc.).
RELLENO_CONECTADO = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FUENTE_CONECTADO = Font(color="006100")
RELLENO_DESCONECTADO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FUENTE_DESCONECTADO = Font(color="9C0006")
RELLENO_SIN_DATO = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FUENTE_SIN_DATO = Font(color="595959")

ESTILOS_STATUS = {
    "Conectado": (RELLENO_CONECTADO, FUENTE_CONECTADO),
    "Desconectado": (RELLENO_DESCONECTADO, FUENTE_DESCONECTADO),
}


MESES_ES = [
    "Ene", "Feb", "Mar", "Abr", "May", "Jun",
    "Jul", "Ago", "Sep", "Oct", "Nov", "Dic",
]


def meses_ordenados(historial):
    """Todos los meses ('YYYY-MM') presentes en el historial, ordenados."""
    return sorted({dia[:7] for dia in dias_ordenados(historial)})


def nombre_hoja_mes(mes):
    anio, num = mes.split("-")
    return f"Resumen {MESES_ES[int(num) - 1]}-{anio}"


def escribir_resumen(nombre_archivo, historial):
    """(Re)crea, en el Excel ya generado, una hoja 'Resumen <Mes>-<Año>' por
    cada mes con datos en el historial: una fila por dispositivo, una
    columna por día del mes, celda coloreada según el status de ese día
    (verde/rojo/gris)."""
    wb = openpyxl.load_workbook(nombre_archivo)
    for nombre in list(wb.sheetnames):
        if nombre.startswith("Resumen "):
            del wb[nombre]

    dias = dias_ordenados(historial)
    ids_ordenados = sorted(historial.keys())

    for mes in meses_ordenados(historial):
        dias_mes = [d for d in dias if d.startswith(mes)]
        ws = wb.create_sheet(nombre_hoja_mes(mes))

        ws.cell(row=1, column=1, value="Ubicación|Entidad").font = Font(bold=True)
        for j, dia in enumerate(dias_mes, start=2):
            celda = ws.cell(row=1, column=j, value=int(dia[-2:]))
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="center")

        for i, id_ in enumerate(ids_ordenados, start=2):
            ws.cell(row=i, column=1, value=id_)
            for j, dia in enumerate(dias_mes, start=2):
                status = historial[id_].get(dia)
                relleno, fuente = ESTILOS_STATUS.get(status, (RELLENO_SIN_DATO, FUENTE_SIN_DATO))
                celda = ws.cell(row=i, column=j, value=status or "")
                celda.fill = relleno
                celda.font = fuente

        ws.column_dimensions["A"].width = 40
        for col in range(2, len(dias_mes) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 4
        ws.freeze_panes = "B2"

    wb.save(nombre_archivo)


def main():
    filas = construir_filas()

    ubicaciones = sorted(set(f["ubicacion"] for f in filas))
    print(f"Ubicaciones encontradas ({len(ubicaciones)}): {', '.join(ubicaciones)}")
    print(f"Total de filas a escribir: {len(filas)}\n")
    for f in filas:
        print(f"  [{f['ubicacion']}] {f['entidad']}: {f['status']}")

    nombre_salida = escribir_excel(filas)

    hoy = datetime.now(TZ_CHILE).strftime("%Y-%m-%d")
    historial = actualizar_historial(filas, hoy)
    escribir_resumen(nombre_salida, historial)

    print(f"\n✅ {len(filas)} filas escritas. Archivo generado: {nombre_salida}")

    with open(os.environ.get("GITHUB_OUTPUT", "/dev/null"), "a") as f:
        f.write(f"archivo={nombre_salida}\n")


if __name__ == "__main__":
    main()
